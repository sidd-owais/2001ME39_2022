from platform import python_version
from datetime import datetime
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
import math
from scipy.stats import rankdata
import os
from io import BytesIO
import base64
import os
import json
import pickle
import uuid
import re

start_time = datetime.now()

# Help


def download_button(object_to_download, download_filename, button_text, pickle_it=False):
    """
    Generates a link to download the given object_to_download.

    Params:
    ------
    object_to_download:  The object to be downloaded.
    download_filename (str): filename and extension of file. e.g. mydata.csv,
    some_txt_output.txt download_link_text (str): Text to display for download
    link.
    button_text (str): Text to display on download button (e.g. 'click here to download file')
    pickle_it (bool): If True, pickle file.

    Returns:
    -------
    (str): the anchor tag to download object_to_download

    Examples:
    --------
    download_link(your_df, 'YOUR_DF.csv', 'Click to download data!')
    download_link(your_str, 'YOUR_STRING.txt', 'Click to download text!')

    """
    # if pickle_it:
    #     try:
    #         object_to_download = pickle.dumps(object_to_download)
    #     except pickle.PicklingError as e:
    #         st.write(e)
    #         return None

    # else:
    #     if isinstance(object_to_download, bytes):
    #         pass

    #     elif isinstance(object_to_download, pd.DataFrame):
    #         object_to_download = object_to_download.to_csv(index=False)

    #     # Try JSON encode for everything else
    #     else:
    #         object_to_download = json.dumps(object_to_download)

    try:
        # some strings <-> bytes conversions necessary here
        b64 = base64.b64encode(object_to_download.encode()).decode()

    except AttributeError as e:
        b64 = base64.b64encode(object_to_download).decode()

    button_uuid = str(uuid.uuid4()).replace('-', '')
    button_id = re.sub('\d+', '', button_uuid)

    custom_css = f""" 
        <style>
            #{button_id} {{
                background-color: rgb(255, 255, 255);
                color: rgb(38, 39, 48);
                padding: 0.25em 0.38em;
                position: relative;
                text-decoration: none;
                border-radius: 4px;
                border-width: 1px;
                border-style: solid;
                border-color: rgb(230, 234, 241);
                border-image: initial;

            }} 
            #{button_id}:hover {{
                border-color: rgb(246, 51, 102);
                color: rgb(246, 51, 102);
            }}
            #{button_id}:active {{
                box-shadow: none;
                background-color: rgb(246, 51, 102);
                color: white;
                }}
        </style> """

    dl_link = custom_css + \
        f'<a download="{download_filename}" id="{button_id}" href="data:file/txt;base64,{b64}">{button_text}</a><br></br>'

    return dl_link


def proj_octant_gui():

    def fun(input, val):
        df = pd.read_excel(input).loc[:1000]
        # mod value
        a = val
        file_name = (input.name).split('.xlsx')[0] + \
            "_octant_analysis_mod_"+str(a)+".xlsx"
        df.to_excel(file_name, index=False, header=True)
        wb = load_workbook(file_name)
        sheet = wb.active
        sheet_name = wb.active.title
        # Creating column for different velocity
        sheet.cell(row=1, column=5).value = "U_avg"
        sheet.cell(row=1, column=6).value = "V_avg"
        sheet.cell(row=1, column=7).value = "W_avg"
        sheet.cell(row=1, column=8).value = "U-U_avg"
        sheet.cell(row=1, column=9).value = "V-V_avg"
        sheet.cell(row=1, column=10).value = "W-W_avg"
        sheet.cell(row=1, column=11).value = "Octant"
        # Average of different coordinate velocity
        U_avg = df["U"].mean()
        V_avg = df["V"].mean()
        W_avg = df["W"].mean()

        # Appending means values to sheet
        sheet.cell(row=2, column=5, value=round(U_avg, 3))
        sheet.cell(row=2, column=6, value=round(V_avg, 3))
        sheet.cell(row=2, column=7, value=round(W_avg, 3))

        # Calculating difference of absolute and mean
        for i in range(2, len(df)+2):
            c2 = sheet.cell(i, 2).value
            sheet.cell(i, 8, round(c2-U_avg, 3))
        for i in range(2, len(df)+2):
            c2 = sheet.cell(i, 3).value
            sheet.cell(i, 9, round(c2-V_avg, 3))
        for i in range(2, len(df)+2):
            c2 = sheet.cell(i, 4).value
            sheet.cell(i, 10, round(c2-W_avg, 3))
        # List for overall count of octant

        octant = [[0 for i in range(9)] for j in range(2)]
        octant[0] = ['Octant ID', '+1', '-1',
                     '+2', '-2', '+3', '-3', '+4', '-4']
        octant[1][0] = "Overall Count"
        octant_sign = [[0] for i in range(len(df))]

        # Deciding octant of different coordinate
        for i in range(2, len(df)+2):
            x = sheet.cell(i, 8).value
            y = sheet.cell(i, 9).value
            z = sheet.cell(i, 10).value
            if (x >= 0 and y >= 0 and z >= 0):
                octant_sign[i-2][0] = "+1"
                octant[1][1] += 1
            if (x < 0 and y >= 0 and z >= 0):
                octant_sign[i-2][0] = "+2"
                octant[1][3] += 1
            if (x < 0 and y < 0 and z >= 0):
                octant_sign[i-2][0] = "+3"
                octant[1][5] += 1
            if (x >= 0 and y < 0 and z >= 0):
                octant_sign[i-2][0] = "+4"
                octant[1][7] += 1
            if (x >= 0 and y >= 0 and z < 0):
                octant_sign[i-2][0] = "-1"
                octant[1][2] += 1
            if (x < 0 and y >= 0 and z < 0):
                octant_sign[i-2][0] = "-2"
                octant[1][4] += 1
            if (x < 0 and y < 0 and z < 0):
                octant_sign[i-2][0] = "-3"
                octant[1][6] += 1
            if (x >= 0 and y < 0 and z < 0):
                octant_sign[i-2][0] = "-4"
                octant[1][8] += 1
        wb.save(file_name)
        # *****************************************************************
        # Appending octant list
        writer = pd.ExcelWriter(file_name,
                                mode='a', if_sheet_exists='overlay')
        oct_1 = pd.DataFrame(octant)
        oct_1.to_excel(writer, sheet_name=sheet_name, startcol=13,
                       startrow=2, index=False, header=False)

        try:
            if (a < 0):
                raise Exception()
        except:
            print("Mod value is negative")
            exit()
        grp = math.ceil(len(df)/a)

        # Creating 2d list for storing number of count of different coordinate in different range
        list_1 = [[0 for i in range(9)] for j in range(grp)]
        list_0 = [str(i*a)+"-"+str((i+1)*a-1) for i in range(grp)]
        for i in range(0, grp):
            if (i == (grp-1)):
                list_1[i][0] = str((i)*a)+"-"+str(len(df)-1)
            else:
                list_1[i][0] = list_0[i]

        # Calculating the number of coordinates in different range
        for i in range(0, grp):
            start = (i)*a+2
            end = (i+1)*a+2
            if (i == (grp-1)):
                end = min(((i+1)*a+2), len(df)+2)
            for j in range(start, end):
                x = sheet.cell(j, 8).value
                y = sheet.cell(j, 9).value
                z = sheet.cell(j, 10).value
                if (x > 0 and y > 0 and z > 0):
                    list_1[i][1] += 1
                if (x < 0 and y > 0 and z > 0):
                    list_1[i][3] += 1
                if (x < 0 and y < 0 and z > 0):
                    list_1[i][5] += 1
                if (x > 0 and y < 0 and z > 0):
                    list_1[i][7] += 1
                if (x > 0 and y > 0 and z < 0):
                    list_1[i][2] += 1
                if (x < 0 and y > 0 and z < 0):
                    list_1[i][4] += 1
                if (x < 0 and y < 0 and z < 0):
                    list_1[i][6] += 1
                if (x > 0 and y < 0 and z < 0):
                    list_1[i][8] += 1

        # printing and appending count of each coordinate in different range
        oct_2 = pd.DataFrame(list_1)
        oct_2.to_excel(writer, sheet_name=sheet_name, startcol=13,
                       startrow=4, index=False, header=False)

        list_2 = [[0 for i in range(9)] for j in range(9)]
        list_2[0] = ['Octant#', '+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
        for i in range(1, 9):
            list_2[i][0] = list_2[0][i]

        # Calculating Transition
        dic_1 = {'+1': 1, '-1': 2, '+2': 3, '-2': 4,
                 '+3': 5, '-3': 6, '+4': 7, '-4': 8}
        for i in range(2, len(df)+2-1):
            x = octant_sign[i-2][0]
            y = octant_sign[i-1][0]
            list_2[dic_1[x]][dic_1[y]] += 1

        # Printing count of each Transition in overall data and appending it
        oct_3 = pd.DataFrame(list_2)
        oct_3.to_excel(writer, sheet_name=sheet_name, startcol=34,
                       startrow=2, index=False, header=False)

        # Calculating Transition in different range of data and appending in xlsx file

        for i in range(grp):
            for l in range(1, 9):
                for k in range(1, 9):
                    list_2[l][k] = 0
            start = (i)*a
            end = (i+1)*a-1
            if (i == grp-1):
                end = min(((i+1)*a-1), len(df)-1)
            for j in range(start, end):
                x = octant_sign[j][0]
                y = octant_sign[j+1][0]
                list_2[dic_1[x]][dic_1[y]] += 1
            tran = pd.DataFrame(list_2)
            tran.to_excel(writer, sheet_name=sheet_name, startcol=34, startrow=16 +
                          (i*14), index=False, header=False)

        # Longest Subsequence count of every octant
        list_00 = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
        dict_2 = {'+1': [], '-1': [], '+2': [], '-2': [],
                  '+3': [], '-3': [], '+4': [], '-4': []}
        list_6 = [[0 for i in range(3)] for j in range(9)]
        list_6[0] = ['Octant##', 'Longest_Subsquence_Length', 'Count']
        for i in range(1, 9):
            list_6[i][0] = list_00[i-1]
        temp_1 = [0]
        start = 0
        for i in range(0, len(df)):
            x = octant_sign[i][0]
            t = sheet.cell(i+2, 1).value
            if i == 2:
                temp_1[0] += 1
            else:
                y = octant_sign[i-1][0]
                if x == y:
                    temp_1[0] += 1
                else:
                    if list_6[dic_1[y]][1] < temp_1[0]:
                        list_6[dic_1[y]][1] = temp_1[0]
                        list_6[dic_1[y]][2] = 1
                        dict_2[y] = [start]
                    elif list_6[dic_1[y]][1] == temp_1[0]:
                        list_6[dic_1[y]][2] += 1
                        dict_2[y].append(start)
                    temp_1[0] = 1
                    start = t
        long_sub = pd.DataFrame(list_6)
        long_sub.to_excel(writer, sheet_name=sheet_name, startcol=44, startrow=2,
                          index=False, header=False)

        # list for storing rank for overall range
        list_3 = [[0 for i in range(8)] for j in range(2)]
        list_3[0] = ['Rank of +1', 'Rank of -1', 'Rank of +2',
                     'Rank of -2', 'Rank of +3', 'Rank of -3', 'Rank of +4', 'Rank of -4']

        # Calculating Rank
        list_3[1] = rankdata(octant[1][1:9], method="dense")
        m_1 = max(list_3[1])+1
        for i in range(8):
            list_3[1][i] = m_1 - list_3[1][i]

        # Appending the rank list
        oct_3 = pd.DataFrame(list_3)
        oct_3.style.to_excel(writer, sheet_name=sheet_name, startcol=22, startrow=2,
                             index=False, header=False)

        # list for storing Rank of data points in different Range

        list_4 = [[0 for j in range(8)] for i in range(grp)]

        # Calculating
        for i in range(grp):
            list_4[i] = rankdata(list_1[i][1:9], method="dense")
            m = max(list_4[i])+1
            for j in range(8):
                list_4[i][j] = m - list_4[i][j]

        # Appending the Rank list
        oct_4 = pd.DataFrame(list_4)
        oct_4.to_excel(writer, sheet_name=sheet_name, startcol=22,
                       startrow=4, index=False, header=False)

        # Calculating Rank1 Octant ID
        Rank1_octant_ID = [[0] for i in range(grp+1)]

        for i in range(8):
            if list_3[1][i] == 1:
                Rank1_octant_ID[0][0] = octant[0][i+1]
                break

        for i in range(grp):
            for j in range(8):
                if list_4[i][j] == 1:
                    Rank1_octant_ID[i+1][0] = octant[0][j+1]
                    break

        oct_6 = pd.DataFrame(Rank1_octant_ID)
        oct_6.to_excel(writer, sheet_name=sheet_name, startcol=30,
                       startrow=3, index=False, header=False)

        # Calculating Rank1 Octant Name
        dict_1 = {'+1': "Internal outward interaction",
                  '-1': "External outward interaction", '+2': "External Ejection", '-2': "Internal Ejection", '+3': "External inward interaction", '-3': "Internal inward interaction", '+4': "Internal sweep", '-4': "External sweep"}
        Rank1_octant_Name = [[0] for i in range(grp+1)]
        for i in range(grp+1):
            Rank1_octant_Name[i][0] = dict_1[Rank1_octant_ID[i][0]]

        oct_7 = pd.DataFrame(Rank1_octant_Name)
        oct_7.to_excel(writer, sheet_name=sheet_name, startcol=31,
                       startrow=3, index=False, header=False)
        # Calculating Count of Rank 1 Mod Values
        list_5 = [[0 for i in range(3)] for j in range(9)]
        list_5[0] = ["Octant ID", "Octant Name", "Count of Rank 1 Mod Values"]
        temp = ["Internal outward interaction", "External outward interaction", "External Ejection", "Internal Ejection",
                "External inward interaction", "Internal inward interaction", "Internal sweep", "External sweep"]
        for i in range(1, 9):
            list_5[i][0] = octant[0][i]
        for i in range(1, 9):
            list_5[i][1] = temp[i-1]

        dict_3 = {"Internal outward interaction": 1, "External outward interaction": 2, "External Ejection": 3, "Internal Ejection": 4,
                  "External inward interaction": 5, "Internal inward interaction": 6, "Internal sweep": 7, "External sweep": 8}

        for i in range(grp+1):
            list_5[dict_3[Rank1_octant_Name[i][0]]][2] += 1

        oct_5 = pd.DataFrame(list_5)
        oct_5.to_excel(writer, sheet_name=sheet_name, startcol=28, startrow=grp +
                       8, index=False, header=False)
        writer.close()
        ########################################################################
        wb = load_workbook(file_name)
        sheet = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        sheet.cell(3, 49, "###octant").border = thin_border
        sheet.cell(3, 50, "Longest_Subsquence_Length").border = thin_border
        sheet.cell(3, 51, "Count").border = thin_border

        row_count = 4
        for i in range(1, 9):
            octant_1 = list_6[i][0]
            long_sub_leng = list_6[i][1]
            count_1 = list_6[i][2]
            sheet.cell(row_count, 49, octant_1).border = thin_border
            sheet.cell(row_count, 50, long_sub_leng).border = thin_border
            sheet.cell(row_count, 51, count_1).border = thin_border
            row_count += 1
            sheet.cell(row_count, 49, "Time").border = thin_border
            sheet.cell(row_count, 50, "From").border = thin_border
            sheet.cell(row_count, 51, "To").border = thin_border
            row_count += 1
            for j in dict_2[octant_1]:
                sheet.cell(row_count, 49).border = thin_border
                sheet.cell(row_count, 50, j).border = thin_border
                sheet.cell(row_count, 51, j+(0.01*(long_sub_leng-1))
                           ).border = thin_border
                row_count += 1

        # Border and coloring

        for i in range(3, 3+(grp+2)):
            for j in range(14, 33):
                sheet.cell(i, j).border = thin_border

        for i in range(grp+9, (grp+18)):
            for j in range(29, 32):
                sheet.cell(i, j).border = thin_border

        temp_2 = [0, 0]
        variable = 0
        for i in range(3, 12):
            for j in range(35, 44):
                sheet.cell(i, j).border = thin_border
                if i != 3 and j != 35 and variable < int(sheet.cell(i, j).value):
                    variable = sheet.cell(i, j).value
                    temp_2[0] = i
                    temp_2[1] = j
            if (i != 3):
                sheet.cell(temp_2[0], temp_2[1]).fill = PatternFill(
                    patternType='solid', fgColor='FFFF00')
            variable = 0

        for i in range(grp):
            start = 17+(i*14)
            for r in range(start, start+9):
                for c in range(35, 44):
                    sheet.cell(r, c).border = thin_border
                    if r != start and c != 35 and variable < int(sheet.cell(r, c).value):
                        variable = sheet.cell(r, c).value
                        temp_2[0] = r
                        temp_2[1] = c
                if (i != start):
                    sheet.cell(temp_2[0], temp_2[1]).fill = PatternFill(
                        patternType='solid', fgColor='FFFF00')
                variable = 0

        for i in range(3, 12):
            for j in range(45, 48):
                sheet.cell(i, j).border = thin_border

        for i in range(4, 4+grp+1):
            for j in range(23, 31):
                if (sheet.cell(i, j).value == 1):
                    sheet.cell(i, j).fill = PatternFill(
                        patternType='solid', fgColor='FFFF00')

        # Formatting

        for i in range(grp):
            start_row = 17 + (i*14)
            col = 35
            sheet.cell(start_row-1, col).value = list_1[i][0]
            sheet.cell(start_row-2, col).value = "Mod Transition Count"
            sheet.cell(start_row+1, col-1).value = "From"

        sheet['AI1'] = "Overall Transition Count"
        sheet['AS1'] = "Longest Subsequence Length"
        sheet['AW1'] = "Longest Subsequence Length with Range"
        sheet['M4'] = "Mod "+str(a)
        sheet['AH4'] = "From"
        sheet['AF3'] = "Rank1 Octant Name"
        sheet['AE3'] = "Rank1 Octant ID"
        sheet['N1'] = "Octant Octant Count"
        data_1 = BytesIO()
        wb.save(data_1)
        filename = (input.name).split('.xlsx')[
            0] + "_octant_analysis_mod_"+str(a)+".xlsx"
        st.text(filename)
        st.download_button(label="Download File", file_name=(input.name).split('.xlsx')[0] +
                           "_octant_analysis_mod_"+str(a)+".xlsx", data=data_1)

    #################################################################################################################################

    def fun_1(input_path, val):
        for file in os.listdir(input_path):
            df = pd.read_excel(input_path+"\\"+file).loc[:1000]
            # mod value
            a = val
            file_name = file.split('.xlsx')[0] + \
                "_octant_analysis_mod_"+str(a)+".xlsx"
            df.to_excel(file_name, index=False, header=True)
            wb = load_workbook(file_name)
            sheet = wb.active
            sheet_name = wb.active.title
            # Creating column for different velocity
            sheet.cell(row=1, column=5).value = "U_avg"
            sheet.cell(row=1, column=6).value = "V_avg"
            sheet.cell(row=1, column=7).value = "W_avg"
            sheet.cell(row=1, column=8).value = "U-U_avg"
            sheet.cell(row=1, column=9).value = "V-V_avg"
            sheet.cell(row=1, column=10).value = "W-W_avg"
            sheet.cell(row=1, column=11).value = "Octant"
            # Average of different coordinate velocity
            U_avg = df["U"].mean()
            V_avg = df["V"].mean()
            W_avg = df["W"].mean()

            # Appending means values to sheet
            sheet.cell(row=2, column=5, value=round(U_avg, 3))
            sheet.cell(row=2, column=6, value=round(V_avg, 3))
            sheet.cell(row=2, column=7, value=round(W_avg, 3))

            # Calculating difference of absolute and mean
            for i in range(2, len(df)+2):
                c2 = sheet.cell(i, 2).value
                sheet.cell(i, 8, round(c2-U_avg, 3))
            for i in range(2, len(df)+2):
                c2 = sheet.cell(i, 3).value
                sheet.cell(i, 9, round(c2-V_avg, 3))
            for i in range(2, len(df)+2):
                c2 = sheet.cell(i, 4).value
                sheet.cell(i, 10, round(c2-W_avg, 3))
            # List for overall count of octant

            octant = [[0 for i in range(9)] for j in range(2)]
            octant[0] = ['Octant ID', '+1', '-1',
                         '+2', '-2', '+3', '-3', '+4', '-4']
            octant[1][0] = "Overall Count"
            octant_sign = [[0] for i in range(len(df))]

            # Deciding octant of different coordinate
            for i in range(2, len(df)+2):
                x = sheet.cell(i, 8).value
                y = sheet.cell(i, 9).value
                z = sheet.cell(i, 10).value
                if (x >= 0 and y >= 0 and z >= 0):
                    octant_sign[i-2][0] = "+1"
                    octant[1][1] += 1
                if (x < 0 and y >= 0 and z >= 0):
                    octant_sign[i-2][0] = "+2"
                    octant[1][3] += 1
                if (x < 0 and y < 0 and z >= 0):
                    octant_sign[i-2][0] = "+3"
                    octant[1][5] += 1
                if (x >= 0 and y < 0 and z >= 0):
                    octant_sign[i-2][0] = "+4"
                    octant[1][7] += 1
                if (x >= 0 and y >= 0 and z < 0):
                    octant_sign[i-2][0] = "-1"
                    octant[1][2] += 1
                if (x < 0 and y >= 0 and z < 0):
                    octant_sign[i-2][0] = "-2"
                    octant[1][4] += 1
                if (x < 0 and y < 0 and z < 0):
                    octant_sign[i-2][0] = "-3"
                    octant[1][6] += 1
                if (x >= 0 and y < 0 and z < 0):
                    octant_sign[i-2][0] = "-4"
                    octant[1][8] += 1
            wb.save(file_name)
            # *****************************************************************************
            # Appending octant list
            writer = pd.ExcelWriter(file_name,
                                    mode='a', if_sheet_exists='overlay')
            oct_1 = pd.DataFrame(octant)
            oct_1.to_excel(writer, sheet_name=sheet_name, startcol=13,
                           startrow=2, index=False, header=False)

            try:
                if (a < 0):
                    raise Exception()
            except:
                print("Mod value is negative")
                exit()
            grp = math.ceil(len(df)/a)

            # Creating 2d list for storing number of count of different coordinate in different range
            list_1 = [[0 for i in range(9)] for j in range(grp)]
            list_0 = [str(i*a)+"-"+str((i+1)*a-1) for i in range(grp)]
            for i in range(0, grp):
                if (i == (grp-1)):
                    list_1[i][0] = str((i)*a)+"-"+str(len(df)-1)
                else:
                    list_1[i][0] = list_0[i]

            # Calculating the number of coordinates in different range
            for i in range(0, grp):
                start = (i)*a+2
                end = (i+1)*a+2
                if (i == (grp-1)):
                    end = min(((i+1)*a+2), len(df)+2)
                for j in range(start, end):
                    x = sheet.cell(j, 8).value
                    y = sheet.cell(j, 9).value
                    z = sheet.cell(j, 10).value
                    if (x > 0 and y > 0 and z > 0):
                        list_1[i][1] += 1
                    if (x < 0 and y > 0 and z > 0):
                        list_1[i][3] += 1
                    if (x < 0 and y < 0 and z > 0):
                        list_1[i][5] += 1
                    if (x > 0 and y < 0 and z > 0):
                        list_1[i][7] += 1
                    if (x > 0 and y > 0 and z < 0):
                        list_1[i][2] += 1
                    if (x < 0 and y > 0 and z < 0):
                        list_1[i][4] += 1
                    if (x < 0 and y < 0 and z < 0):
                        list_1[i][6] += 1
                    if (x > 0 and y < 0 and z < 0):
                        list_1[i][8] += 1

            # printing and appending count of each coordinate in different range
            oct_2 = pd.DataFrame(list_1)
            oct_2.to_excel(writer, sheet_name=sheet_name, startcol=13,
                           startrow=4, index=False, header=False)

            list_2 = [[0 for i in range(9)] for j in range(9)]
            list_2[0] = ['Octant#', '+1', '-1',
                         '+2', '-2', '+3', '-3', '+4', '-4']
            for i in range(1, 9):
                list_2[i][0] = list_2[0][i]

            # Calculating Transition
            dic_1 = {'+1': 1, '-1': 2, '+2': 3, '-2': 4,
                     '+3': 5, '-3': 6, '+4': 7, '-4': 8}
            for i in range(2, len(df)+2-1):
                x = octant_sign[i-2][0]
                y = octant_sign[i-1][0]
                list_2[dic_1[x]][dic_1[y]] += 1

            # Printing count of each Transition in overall data and appending it
            oct_3 = pd.DataFrame(list_2)
            oct_3.to_excel(writer, sheet_name=sheet_name, startcol=34,
                           startrow=2, index=False, header=False)

            # Calculating Transition in different range of data and appending in xlsx file

            for i in range(grp):
                for l in range(1, 9):
                    for k in range(1, 9):
                        list_2[l][k] = 0
                start = (i)*a
                end = (i+1)*a-1
                if (i == grp-1):
                    end = min(((i+1)*a-1), len(df)-1)
                for j in range(start, end):
                    x = octant_sign[j][0]
                    y = octant_sign[j+1][0]
                    list_2[dic_1[x]][dic_1[y]] += 1
                tran = pd.DataFrame(list_2)
                tran.to_excel(writer, sheet_name=sheet_name, startcol=34, startrow=16 +
                              (i*14), index=False, header=False)

            # Longest Subsequence count of every octant
            list_00 = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
            dict_2 = {'+1': [], '-1': [], '+2': [], '-2': [],
                      '+3': [], '-3': [], '+4': [], '-4': []}
            list_6 = [[0 for i in range(3)] for j in range(9)]
            list_6[0] = ['Octant##', 'Longest_Subsquence_Length', 'Count']
            for i in range(1, 9):
                list_6[i][0] = list_00[i-1]
            temp_1 = [0]
            start = 0
            for i in range(0, len(df)):
                x = octant_sign[i][0]
                t = sheet.cell(i+2, 1).value
                if i == 2:
                    temp_1[0] += 1
                else:
                    y = octant_sign[i-1][0]
                    if x == y:
                        temp_1[0] += 1
                    else:
                        if list_6[dic_1[y]][1] < temp_1[0]:
                            list_6[dic_1[y]][1] = temp_1[0]
                            list_6[dic_1[y]][2] = 1
                            dict_2[y] = [start]
                        elif list_6[dic_1[y]][1] == temp_1[0]:
                            list_6[dic_1[y]][2] += 1
                            dict_2[y].append(start)
                        temp_1[0] = 1
                        start = t
            long_sub = pd.DataFrame(list_6)
            long_sub.to_excel(writer, sheet_name=sheet_name, startcol=44, startrow=2,
                              index=False, header=False)

            # list for storing rank for overall range
            list_3 = [[0 for i in range(8)] for j in range(2)]
            list_3[0] = ['Rank of +1', 'Rank of -1', 'Rank of +2',
                         'Rank of -2', 'Rank of +3', 'Rank of -3', 'Rank of +4', 'Rank of -4']

            # Calculating Rank
            list_3[1] = rankdata(octant[1][1:9], method="dense")
            m_1 = max(list_3[1])+1
            for i in range(8):
                list_3[1][i] = m_1 - list_3[1][i]

            # Appending the rank list
            oct_3 = pd.DataFrame(list_3)
            oct_3.style.to_excel(writer, sheet_name=sheet_name, startcol=22, startrow=2,
                                 index=False, header=False)

            # list for storing Rank of data points in different Range

            list_4 = [[0 for j in range(8)] for i in range(grp)]

            # Calculating
            for i in range(grp):
                list_4[i] = rankdata(list_1[i][1:9], method="dense")
                m = max(list_4[i])+1
                for j in range(8):
                    list_4[i][j] = m - list_4[i][j]

            # Appending the Rank list
            oct_4 = pd.DataFrame(list_4)
            oct_4.to_excel(writer, sheet_name=sheet_name, startcol=22,
                           startrow=4, index=False, header=False)

            # Calculating Rank1 Octant ID
            Rank1_octant_ID = [[0] for i in range(grp+1)]

            for i in range(8):
                if list_3[1][i] == 1:
                    Rank1_octant_ID[0][0] = octant[0][i+1]
                    break

            for i in range(grp):
                for j in range(8):
                    if list_4[i][j] == 1:
                        Rank1_octant_ID[i+1][0] = octant[0][j+1]
                        break

            oct_6 = pd.DataFrame(Rank1_octant_ID)
            oct_6.to_excel(writer, sheet_name=sheet_name, startcol=30,
                           startrow=3, index=False, header=False)

            # Calculating Rank1 Octant Name
            dict_1 = {'+1': "Internal outward interaction",
                      '-1': "External outward interaction", '+2': "External Ejection", '-2': "Internal Ejection", '+3': "External inward interaction", '-3': "Internal inward interaction", '+4': "Internal sweep", '-4': "External sweep"}
            Rank1_octant_Name = [[0] for i in range(grp+1)]
            for i in range(grp+1):
                Rank1_octant_Name[i][0] = dict_1[Rank1_octant_ID[i][0]]

            oct_7 = pd.DataFrame(Rank1_octant_Name)
            oct_7.to_excel(writer, sheet_name=sheet_name, startcol=31,
                           startrow=3, index=False, header=False)
            # Calculating Count of Rank 1 Mod Values
            list_5 = [[0 for i in range(3)] for j in range(9)]
            list_5[0] = ["Octant ID", "Octant Name",
                         "Count of Rank 1 Mod Values"]
            temp = ["Internal outward interaction", "External outward interaction", "External Ejection", "Internal Ejection",
                    "External inward interaction", "Internal inward interaction", "Internal sweep", "External sweep"]
            for i in range(1, 9):
                list_5[i][0] = octant[0][i]
            for i in range(1, 9):
                list_5[i][1] = temp[i-1]

            dict_3 = {"Internal outward interaction": 1, "External outward interaction": 2, "External Ejection": 3, "Internal Ejection": 4,
                      "External inward interaction": 5, "Internal inward interaction": 6, "Internal sweep": 7, "External sweep": 8}

            for i in range(grp+1):
                list_5[dict_3[Rank1_octant_Name[i][0]]][2] += 1

            oct_5 = pd.DataFrame(list_5)
            oct_5.to_excel(writer, sheet_name=sheet_name, startcol=28, startrow=grp +
                           8, index=False, header=False)
            writer.close()
            ########################################################################
            wb = load_workbook(file_name)
            sheet = wb.active
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            sheet.cell(3, 49, "###octant").border = thin_border
            sheet.cell(3, 50, "Longest_Subsquence_Length").border = thin_border
            sheet.cell(3, 51, "Count").border = thin_border

            row_count = 4
            for i in range(1, 9):
                octant_1 = list_6[i][0]
                long_sub_leng = list_6[i][1]
                count_1 = list_6[i][2]
                sheet.cell(row_count, 49, octant_1).border = thin_border
                sheet.cell(row_count, 50, long_sub_leng).border = thin_border
                sheet.cell(row_count, 51, count_1).border = thin_border
                row_count += 1
                sheet.cell(row_count, 49, "Time").border = thin_border
                sheet.cell(row_count, 50, "From").border = thin_border
                sheet.cell(row_count, 51, "To").border = thin_border
                row_count += 1
                for j in dict_2[octant_1]:
                    sheet.cell(row_count, 49).border = thin_border
                    sheet.cell(row_count, 50, j).border = thin_border
                    sheet.cell(row_count, 51, j+(0.01*(long_sub_leng-1))
                               ).border = thin_border
                    row_count += 1

            # Border and coloring

            for i in range(3, 3+(grp+2)):
                for j in range(14, 33):
                    sheet.cell(i, j).border = thin_border

            for i in range(grp+9, (grp+18)):
                for j in range(29, 32):
                    sheet.cell(i, j).border = thin_border

            temp_2 = [0, 0]
            variable = 0
            for i in range(3, 12):
                for j in range(35, 44):
                    sheet.cell(i, j).border = thin_border
                    if i != 3 and j != 35 and variable < int(sheet.cell(i, j).value):
                        variable = sheet.cell(i, j).value
                        temp_2[0] = i
                        temp_2[1] = j
                if (i != 3):
                    sheet.cell(temp_2[0], temp_2[1]).fill = PatternFill(
                        patternType='solid', fgColor='FFFF00')
                variable = 0

            for i in range(grp):
                start = 17+(i*14)
                for r in range(start, start+9):
                    for c in range(35, 44):
                        sheet.cell(r, c).border = thin_border
                        if r != start and c != 35 and variable < int(sheet.cell(r, c).value):
                            variable = sheet.cell(r, c).value
                            temp_2[0] = r
                            temp_2[1] = c
                    if (i != start):
                        sheet.cell(temp_2[0], temp_2[1]).fill = PatternFill(
                            patternType='solid', fgColor='FFFF00')
                    variable = 0

            for i in range(3, 12):
                for j in range(45, 48):
                    sheet.cell(i, j).border = thin_border

            for i in range(4, 4+grp+1):
                for j in range(23, 31):
                    if (sheet.cell(i, j).value == 1):
                        sheet.cell(i, j).fill = PatternFill(
                            patternType='solid', fgColor='FFFF00')

            # Formatting

            for i in range(grp):
                start_row = 17 + (i*14)
                col = 35
                sheet.cell(start_row-1, col).value = list_1[i][0]
                sheet.cell(start_row-2, col).value = "Mod Transition Count"
                sheet.cell(start_row+1, col-1).value = "From"

            sheet['AI1'] = "Overall Transition Count"
            sheet['AS1'] = "Longest Subsequence Length"
            sheet['AW1'] = "Longest Subsequence Length with Range"
            sheet['M4'] = "Mod "+str(a)
            sheet['AH4'] = "From"
            sheet['AF3'] = "Rank1 Octant Name"
            sheet['AE3'] = "Rank1 Octant ID"
            sheet['N1'] = "Octant Octant Count"
            data_1 = BytesIO()
            wb.save(data_1)
            filename = file.split('.xlsx')[
                0] + "_octant_analysis_mod_"+str(a)+".xlsx"
            st.text(filename)
            download_button_str = download_button(data_1.getvalue(), download_filename=file.split('.xlsx')[0] +
                                                  "_octant_analysis_mod_"+str(a)+".xlsx", button_text="Click here to download")
            st.markdown(download_button_str, unsafe_allow_html=True)
    st.header("Project - 2")
    val = st.number_input(label="Enter Mod Value", min_value=1, step=1)
    file = st.file_uploader("Choose a File", type=["xlsx"])
    if st.button("Compute"):
        if file == None:
            st.write("Please select a file")
        else:
            fun(file, val)
    direc = st.text_input(label="Enter Path of Directory")
    if st.button("Compute_all_file"):
        if direc == "":
            st.write("Please give a path")
        else:
            fun_1(direc, val)


# Code

ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


proj_octant_gui()


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
