from platform import python_version
from datetime import datetime
start_time = datetime.now()

# Help https://youtu.be/N6PBd4XdnEw


def octant_range_names(mod=5000):

    octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                              "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill
    import math
    from scipy.stats import rankdata
    df = pd.read_excel("octant_input.xlsx")

    # Using try and except
    try:
        wb = load_workbook("octant_input.xlsx")
    except:
        print("Input file missing")
        exit()
    sheet = wb.active

    # Creating column for different velocity
    sheet.cell(row=1, column=5).value = "U_avg"
    sheet.cell(row=1, column=6).value = "V_avg"
    sheet.cell(row=1, column=7).value = "W_avg"
    sheet.cell(row=1, column=8).value = "U-U_avg"
    sheet.cell(row=1, column=9).value = "V-V_avg"
    sheet.cell(row=1, column=10).value = "W-W_avg"
    sheet.cell(row=1, column=11).value = "Octant"

    # Average of different coordinate velocity
    U_avg = df["U"].mean()
    V_avg = df["V"].mean()
    W_avg = df["W"].mean()

    # Appending means values to sheet
    sheet.cell(row=2, column=5, value=U_avg)
    sheet.cell(row=2, column=6, value=V_avg)
    sheet.cell(row=2, column=7, value=W_avg)

    # Calculating difference of absolute and mean
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 2).value
        sheet.cell(i, 8, c2-U_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 3).value
        sheet.cell(i, 9, c2-V_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 4).value
        sheet.cell(i, 10, c2-W_avg)

    # List for overall count of octant
    octant = [[0 for i in range(8)] for j in range(2)]
    octant[0] = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    # Deciding octant of different coordinate
    for i in range(2, len(df)+2):
        x = sheet.cell(i, 8).value
        y = sheet.cell(i, 9).value
        z = sheet.cell(i, 10).value
        if (x > 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+1")
            octant[1][0] += 1
        if (x < 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+2")
            octant[1][2] += 1
        if (x < 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+3")
            octant[1][4] += 1
        if (x > 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+4")
            octant[1][6] += 1
        if (x > 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-1")
            octant[1][1] += 1
        if (x < 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-2")
            octant[1][3] += 1
        if (x < 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-3")
            octant[1][5] += 1
        if (x > 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-4")
            octant[1][7] += 1
    wb.save("octant_output_ranking_excel.xlsx")
    # Appending octant list
    oct_1 = pd.DataFrame(octant)
    writer = pd.ExcelWriter('octant_output_ranking_excel.xlsx',
                            mode='a', if_sheet_exists='overlay')
    oct_1.to_excel(writer, startcol=13, startrow=1, index=False, header=False)

    # mod value
    a = mod
    try:
        if (a < 0):
            raise Exception()
    except:
        print("Mod value is negative")
        exit()
    grp = math.ceil(len(df)/a)
    # Try and except
    try:
        wb = load_workbook("octant_output_ranking_excel.xlsx")
    except:
        print("Output file missing")
        exit()
    sheet = wb.active

    # Creating 2d list for storing number of count of different coordinate in different range
    list_1 = [[0 for i in range(9)] for j in range(grp+1)]
    list_0 = [str(i*a)+"-"+str((i+1)*a-1) for i in range(grp)]
    list_1[0] = ['Mod 5000', '+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    for i in range(1, grp+1):
        if (i == (grp)):
            list_1[i][0] = str((i-1)*a)+"-"+str(len(df)-1)
        else:
            list_1[i][0] = list_0[i-1]

    # Calculating the number of coordinates in different range
    for i in range(1, grp+1):
        start = (i-1)*a+2
        end = (i)*a+2
        if (i == grp):
            end = min(((i)*a+2), len(df)+2)
        for j in range(start, end):
            x = sheet.cell(j, 8).value
            y = sheet.cell(j, 9).value
            z = sheet.cell(j, 10).value
            if (x > 0 and y > 0 and z > 0):
                list_1[i][1] += 1
            if (x < 0 and y > 0 and z > 0):
                list_1[i][3] += 1
            if (x < 0 and y < 0 and z > 0):
                list_1[i][5] += 1
            if (x > 0 and y < 0 and z > 0):
                list_1[i][7] += 1
            if (x > 0 and y > 0 and z < 0):
                list_1[i][2] += 1
            if (x < 0 and y > 0 and z < 0):
                list_1[i][4] += 1
            if (x < 0 and y < 0 and z < 0):
                list_1[i][6] += 1
            if (x > 0 and y < 0 and z < 0):
                list_1[i][8] += 1

    # printing and appending count of each coordinate in different range
    oct_2 = pd.DataFrame(list_1)
    oct_2.to_excel(writer, startcol=12, startrow=3, index=False, header=False)
    # list for storing rank for overall range
    list_3 = [[0 for i in range(8)] for j in range(3)]
    list_3[0] = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    list_3[1] = ['Rank1', 'Rank2', 'Rank3',
                 'Rank4', 'Rank5', 'Rank6', 'Rank7', 'Rank8']

    # Calculating Rank
    list_3[2] = rankdata(octant[1])
    for i in range(8):
        list_3[2][i] = 9 - list_3[2][i]

    # Appending the rank list
    oct_3 = pd.DataFrame(list_3)
    oct_3.to_excel(writer, startcol=21, startrow=0, index=False, header=False)
    # list for storing Rank of data points in different Range
    dict_1 = {'+1': "Internal outward interaction",
              '-1': "External outward interaction", '+2': "External Ejection", '-2': "Internal Ejection", '+3': "External inward interaction", '-3': "Internal inward interaction", '+4': "Internal sweep", '-4': "External sweep"}
    list_4 = [[0 for j in range(8)] for i in range(grp)]

    # Calculating
    for i in range(grp):
        list_4[i] = rankdata(list_1[i+1][1:9])
        for j in range(8):
            list_4[i][j] = 9 - list_4[i][j]
    # Appending the Rank list
    oct_4 = pd.DataFrame(list_4)
    oct_4.to_excel(writer, startcol=21, startrow=4, index=False, header=False)
    writer.close()
    # Calculating Rank1 Octant ID
    wb = load_workbook("octant_output_ranking_excel.xlsx")
    sheet = wb.active
    sheet['AD2'].value = "Rank1 Octant ID"
    for i in range(8):
        if list_3[2][i] == 1:
            sheet['AD3'].value = octant[0][i]
            break

    count = 5

    for i in range(grp):
        for j in range(8):
            if list_4[i][j] == 1:
                sheet.cell(row=count, column=30).value = octant[0][j]
                count += 1
                break
    # Calculating Rank1 Octant Name
    sheet['AE2'].value = "Rank1 Octant Name"
    sheet['AE3'].value = dict_1[sheet['AD3'].value]
    count = 5
    for i in range(grp):
        ans = sheet.cell(row=count, column=30).value
        sheet.cell(row=count, column=31).value = dict_1[ans]
        count += 1
    sheet['M2'].value = "Octant ID"
    sheet['M3'].value = "Overall Count"
    sheet['L4'].value = "User Input"
    wb.save("octant_output_ranking_excel.xlsx")
    # Calculating Count of Rank 1 Mod Values
    writer = pd.ExcelWriter('octant_output_ranking_excel.xlsx',
                            mode='a', if_sheet_exists='overlay')
    list_5 = [[0 for i in range(3)] for j in range(9)]
    list_5[0] = ["Octant ID", "Octant Name", "Count of Rank 1 Mod Values"]
    temp = ["Internal outward interaction", "External outward interaction", "External Ejection", "Internal Ejection",
            "External inward interaction", "Internal inward interaction", "Internal sweep", "External sweep"]
    for i in range(1, 9):
        list_5[i][0] = list_3[0][i-1]
    for i in range(1, 9):
        list_5[i][1] = temp[i-1]
    count = 5
    dict_2 = {"Internal outward interaction": 1, "External outward interaction": 2, "External Ejection": 3, "Internal Ejection": 4,
              "External inward interaction": 5, "Internal inward interaction": 6, "Internal sweep": 7, "External sweep": 8}
    for i in range(grp):
        ans = sheet.cell(row=count, column=31).value
        list_5[dict_2[ans]][2] += 1
        count += 1

    oct_5 = pd.DataFrame(list_5)
    oct_5.to_excel(writer, startcol=13, startrow=grp +
                   8, index=False, header=False)

    writer.close()


ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod = 5000
octant_range_names(mod)


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
