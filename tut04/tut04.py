from platform import python_version
from datetime import datetime
start_time = datetime.now()


def octant_longest_subsequence_count_with_range():
    import pandas as pd
    from openpyxl import load_workbook
    import math
    df = pd.read_excel("input_octant_longest_subsequence.xlsx")
    # Using try and except
    try:
        wb = load_workbook("input_octant_longest_subsequence.xlsx")
    except:
        print("Input file missing")

    sheet = wb.active

    # Creating column for different velocity
    sheet.cell(row=1, column=5).value = "U_avg"
    sheet.cell(row=1, column=6).value = "V_avg"
    sheet.cell(row=1, column=7).value = "W_avg"
    sheet.cell(row=1, column=8).value = "U-U_avg"
    sheet.cell(row=1, column=9).value = "V-V_avg"
    sheet.cell(row=1, column=10).value = "W-W_avg"
    sheet.cell(row=1, column=11).value = "Octant"

    # Average of different coordinate velocity
    U_avg = df["U"].mean()
    V_avg = df["V"].mean()
    W_avg = df["W"].mean()

    # Appending means values to sheet
    sheet.cell(row=2, column=5, value=U_avg)
    sheet.cell(row=2, column=6, value=V_avg)
    sheet.cell(row=2, column=7, value=W_avg)

    # Calculating difference of absolute and mean
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 2).value
        sheet.cell(i, 8, c2-U_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 3).value
        sheet.cell(i, 9, c2-V_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 4).value
        sheet.cell(i, 10, c2-W_avg)

    # Deciding octant of different coordinate
    for i in range(2, len(df)+2):
        x = sheet.cell(i, 8).value
        y = sheet.cell(i, 9).value
        z = sheet.cell(i, 10).value
        if (x > 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+1")
        if (x < 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+2")
        if (x < 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+3")
        if (x > 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+4")
        if (x > 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-1")
        if (x < 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-2")
        if (x < 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-3")
        if (x > 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-4")
    wb.save("output.xlsx")
    # Longest Subsequence count of every octant and their time interval
    list_1 = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    dic_1 = {'+1': 1, '-1': 2, '+2': 3, '-2': 4,
             '+3': 5, '-3': 6, '+4': 7, '-4': 8}
    list_2 = [[0 for i in range(3)] for j in range(9)]
    list_2[0] = ['Octant', 'Longest_Subsquence_Length', 'Count']

    writer = pd.ExcelWriter('output.xlsx', mode='a', if_sheet_exists='overlay')

    for i in range(1, 9):
        list_2[i][0] = list_1[i-1]
    dict_2 = {'+1': [], '-1': [], '+2': [], '-2': [],
              '+3': [], '-3': [], '+4': [], '-4': []}

    start = 0
    temp = [0]
    for i in range(2, len(df)+2):
        x = sheet.cell(i, 11).value
        t = sheet.cell(i, 1).value
        if i == 2:
            temp[0] += 1
        else:
            y = sheet.cell(i-1, 11).value
            if x == y:
                temp[0] += 1
            else:
                if list_2[dic_1[y]][1] < temp[0]:
                    list_2[dic_1[y]][1] = temp[0]
                    list_2[dic_1[y]][2] = 1
                    dict_2[y] = [start]
                elif list_2[dic_1[y]][1] == temp[0]:
                    list_2[dic_1[y]][2] += 1
                    dict_2[y].append(start)
                temp[0] = 1
                start = t

    long_sub = pd.DataFrame(list_2)
    long_sub.to_excel(writer, startcol=12, startrow=0,
                      index=False, header=False)
    writer.close()
    # Writing in excel sheet

    wb = load_workbook("output.xlsx")
    sheet = wb.active

    sheet.cell(1, 17, "octant")
    sheet.cell(1, 18, "Longest_Subsquence_Length")
    sheet.cell(1, 19, "Count")

    row_count = 2
    for i in range(1, 9):
        octant = list_2[i][0]
        long_sub_leng = list_2[i][1]
        count_1 = list_2[i][2]
        sheet.cell(row_count, 17, octant)
        sheet.cell(row_count, 18, long_sub_leng)
        sheet.cell(row_count, 19, count_1)
        row_count += 1
        sheet.cell(row_count, 17, "Time")
        sheet.cell(row_count, 18, "From")
        sheet.cell(row_count, 19, "To")
        row_count += 1
        for j in dict_2[octant]:
            sheet.cell(row_count, 18, j)
            sheet.cell(row_count, 19, j+(0.01*(long_sub_leng-1)))
            row_count += 1

    wb.save("output.xlsx")
