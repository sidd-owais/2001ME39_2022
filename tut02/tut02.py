from platform import python_version


def octant_transition_count(mod=5000):
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill
    import math
    df = pd.read_excel("input_octant_transition_identify.xlsx")

    # Using try and except
    try:
        wb = load_workbook("input_octant_transition_identify.xlsx")
    except:
        print("Input file missing")
        exit()
    sheet = wb.active

    # Creating column for different velocity
    sheet.cell(row=1, column=5).value = "U_avg"
    sheet.cell(row=1, column=6).value = "V_avg"
    sheet.cell(row=1, column=7).value = "W_avg"
    sheet.cell(row=1, column=8).value = "U-U_avg"
    sheet.cell(row=1, column=9).value = "V-V_avg"
    sheet.cell(row=1, column=10).value = "W-W_avg"
    sheet.cell(row=1, column=11).value = "Octant"

    # Average of different coordinate velocity
    U_avg = df["U"].mean()
    V_avg = df["V"].mean()
    W_avg = df["W"].mean()

    # Appending means values to sheet
    sheet.cell(row=2, column=5, value=U_avg)
    sheet.cell(row=2, column=6, value=V_avg)
    sheet.cell(row=2, column=7, value=W_avg)

    # Calculating difference of absolute and mean
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 2).value
        sheet.cell(i, 8, c2-U_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 3).value
        sheet.cell(i, 9, c2-V_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 4).value
        sheet.cell(i, 10, c2-W_avg)

    # Deciding octant of different coordinate
    octant = [0 for i in range(8)]  # List for overall count of octant
    for i in range(2, len(df)+2):
        x = sheet.cell(i, 8).value
        y = sheet.cell(i, 9).value
        z = sheet.cell(i, 10).value
        if (x > 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+1")
            octant[0] += 1
        if (x < 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+2")
            octant[2] += 1
        if (x < 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+3")
            octant[4] += 1
        if (x > 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+4")
            octant[6] += 1
        if (x > 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-1")
            octant[1] += 1
        if (x < 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-2")
            octant[3] += 1
        if (x < 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-3")
            octant[5] += 1
        if (x > 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-4")
            octant[7] += 1
    wb.save("output_octant_transition_identify.xlsx")

    # Appending octant list
    temp = [[0 for i in range(8)] for i in range(1)]
    oct = pd.DataFrame(
        temp, columns=['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4'])
    oct.iloc[0] = octant
    writer = pd.ExcelWriter(
        'output_octant_transition_identify.xlsx', mode='a', if_sheet_exists='overlay')
    oct.to_excel(writer, startcol=13, startrow=0, index=False)

    # mod value
    a = mod
    try:
        if (a < 0):
            raise Exception()
    except:
        print("Mod value is negative")
        exit()
    grp = math.ceil(len(df)/a)
    # Try and except
    try:
        wb = load_workbook("output_octant_transition_identify.xlsx")
    except:
        print("Output file missing")
        exit()
    sheet = wb.active

    # Creating 2d list for storing number of count of different coordinate in different range
    list_1 = [[0 for i in range(9)] for j in range(grp+1)]
    list_0 = [str(i*a)+"-"+str((i+1)*a-1) for i in range(grp)]
    list_1[0] = ['Mod 5000', '+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    for i in range(1, grp+1):
        if (i == (grp)):
            list_1[i][0] = str((i-1)*a)+"-"+str(len(df)-1)
        else:
            list_1[i][0] = list_0[i-1]

    # Calculating the number of coordinates in different range
    for i in range(1, grp+1):
        start = (i-1)*a+2
        end = (i)*a+2
        if (i == grp):
            end = min(((i)*a+2), len(df)+2)
        for j in range(start, end):
            x = sheet.cell(j, 8).value
            y = sheet.cell(j, 9).value
            z = sheet.cell(j, 10).value
            if (x > 0 and y > 0 and z > 0):
                list_1[i][1] += 1
            if (x < 0 and y > 0 and z > 0):
                list_1[i][3] += 1
            if (x < 0 and y < 0 and z > 0):
                list_1[i][5] += 1
            if (x > 0 and y < 0 and z > 0):
                list_1[i][7] += 1
            if (x > 0 and y > 0 and z < 0):
                list_1[i][2] += 1
            if (x < 0 and y > 0 and z < 0):
                list_1[i][4] += 1
            if (x < 0 and y < 0 and z < 0):
                list_1[i][6] += 1
            if (x > 0 and y < 0 and z < 0):
                list_1[i][8] += 1

    # printing and appending count of each coordinate in different range
    oct_2 = pd.DataFrame(list_1)
    oct_2.to_excel(writer, startcol=12, startrow=4, index=False, header=False)
    # Transition
    list_2 = [[0 for i in range(9)] for j in range(9)]
    list_2[0] = ['Count', '+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    for i in range(1, 9):
        list_2[i][0] = list_2[0][i]

    # Calculating Transition
    dic_1 = {'+1': 1, '-1': 2, '+2': 3, '-2': 4,
             '+3': 5, '-3': 6, '+4': 7, '-4': 8}
    for i in range(2, len(df)+2-1):
        x = sheet.cell(i, 11).value
        y = sheet.cell(i+1, 11).value
        list_2[dic_1[x]][dic_1[y]] += 1

    # Printing count of each Transition in overall data and appending it
    oct_3 = pd.DataFrame(list_2)
    oct_3.to_excel(writer, startcol=12, startrow=15, index=False, header=False)

    # Calculating Transition in different range of data and appending in xlsx file
    for i in range(grp):
        for l in range(1, 9):
            for k in range(1, 9):
                list_2[l][k] = 0
        start = (i)*a+2
        end = (i+1)*a+2
        if (i == grp-1):
            end = min(((i+1)*a+2), len(df)+2-1)
        for j in range(start, end):
            x = sheet.cell(j, 11).value
            y = sheet.cell(j+1, 11).value
            list_2[dic_1[x]][dic_1[y]] += 1
        tran = pd.DataFrame(list_2)
        tran.to_excel(writer, startcol=12, startrow=28 +
                      (i*11), index=False, header=False)
    writer.close()
    # Colouring the particular cell
    wb = load_workbook("output_octant_transition_identify.xlsx")
    sheet = wb.active
    sheet['L5'].value = "User Input"
    sheet['L5'].fill = PatternFill(patternType='solid', fgColor='00FFFF00')
    sheet['M5'].fill = PatternFill(patternType='solid', fgColor='F0FF0F0F')
    # Heading of different range and little bit of formatting
    for i in range(grp):
        if i == grp-1:
            sheet.cell(row=29+(i*11), column=13).value = str((i)*a) + \
                "-"+str(len(df)-1)
        else:
            sheet.cell(row=29+(i*11), column=13).value = list_0[i]
    for i in range(grp):
        sheet.cell(row=28+(i*11), column=14).value = "To"
    for i in range(grp):
        sheet.cell(row=30+(i*11), column=12).value = "From"
    sheet['M2'].value = "Overall Count"
    sheet['M14'].value = "Overall Transition count"
    sheet['N15'] = "To"
    sheet['L17'] = "From"
    wb.save('output_octant_transition_identify.xlsx')


ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

mod = 5000
# Calling function
octant_transition_count(mod)
