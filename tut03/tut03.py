from platform import python_version
from datetime import datetime
start_time = datetime.now()


def octant_longest_subsequence_count():
    import pandas as pd
    from openpyxl import load_workbook
    import math
    df = pd.read_excel("input_octant_longest_subsequence.xlsx")
    # Using try and except
    try:
        wb = load_workbook("input_octant_longest_subsequence.xlsx")
    except:
        print("Input file missing")

    sheet = wb.active

    # Creating column for different velocity
    sheet.cell(row=1, column=5).value = "U_avg"
    sheet.cell(row=1, column=6).value = "V_avg"
    sheet.cell(row=1, column=7).value = "W_avg"
    sheet.cell(row=1, column=8).value = "U-U_avg"
    sheet.cell(row=1, column=9).value = "V-V_avg"
    sheet.cell(row=1, column=10).value = "W-W_avg"
    sheet.cell(row=1, column=11).value = "Octant"

    # Average of different coordinate velocity
    U_avg = df["U"].mean()
    V_avg = df["V"].mean()
    W_avg = df["W"].mean()

    # Appending means values to sheet
    sheet.cell(row=2, column=5, value=U_avg)
    sheet.cell(row=2, column=6, value=V_avg)
    sheet.cell(row=2, column=7, value=W_avg)

    # Calculating difference of absolute and mean
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 2).value
        sheet.cell(i, 8, c2-U_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 3).value
        sheet.cell(i, 9, c2-V_avg)
    for i in range(2, len(df)+2):
        c2 = sheet.cell(i, 4).value
        sheet.cell(i, 10, c2-W_avg)

    # Deciding octant of different coordinate
    for i in range(2, len(df)+2):
        x = sheet.cell(i, 8).value
        y = sheet.cell(i, 9).value
        z = sheet.cell(i, 10).value
        if (x > 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+1")
        if (x < 0 and y > 0 and z > 0):
            sheet.cell(i, 11, "+2")
        if (x < 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+3")
        if (x > 0 and y < 0 and z > 0):
            sheet.cell(i, 11, "+4")
        if (x > 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-1")
        if (x < 0 and y > 0 and z < 0):
            sheet.cell(i, 11, "-2")
        if (x < 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-3")
        if (x > 0 and y < 0 and z < 0):
            sheet.cell(i, 11, "-4")
    wb.save("output_octant_longest_subsequence.xlsx")

    # Longest Subsequence count of every octant
    list_1 = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']
    dic_1 = {'+1': 1, '-1': 2, '+2': 3, '-2': 4,
             '+3': 5, '-3': 6, '+4': 7, '-4': 8}
    list_2 = [[0 for i in range(3)] for j in range(9)]
    list_2[0] = ['Octant', 'Longest_Subsquence_Length', 'Count']
    writer = pd.ExcelWriter(
        'output_octant_longest_subsequence.xlsx', mode='a', if_sheet_exists='overlay')
    for i in range(1, 9):
        list_2[i][0] = list_1[i-1]
    temp = [0]
    for i in range(2, len(df)+2):
        x = sheet.cell(i, 11).value
        if i == 2:
            temp[0] += 1
        else:
            y = sheet.cell(i-1, 11).value
            if x == y:
                temp[0] += 1
            else:
                if list_2[dic_1[y]][1] < temp[0]:
                    list_2[dic_1[y]][1] = temp[0]
                    list_2[dic_1[y]][2] = 1
                elif list_2[dic_1[y]][1] == temp[0]:
                    list_2[dic_1[y]][2] += 1
                temp[0] = 1
    long_sub = pd.DataFrame(list_2)
    long_sub.to_excel(writer, startcol=12, startrow=0,
                      index=False, header=False)
    writer.close()


ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


octant_longest_subsequence_count()


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
